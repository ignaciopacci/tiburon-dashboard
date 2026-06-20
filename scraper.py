import os
import time
import json
import re
import requests
from datetime import datetime, timedelta, timezone
import calendar
import pdfplumber
from io import BytesIO
from playwright.sync_api import sync_playwright
import dropbox
import base64
from google.oauth2 import service_account
from googleapiclient.discovery import build
from googleapiclient.http import MediaIoBaseDownload
import io
import openpyxl

ARG = timezone(timedelta(hours=-3))

USUARIO = os.environ['AMS_USUARIO']
PASSWORD = os.environ['AMS_PASSWORD']
EMPRESA_1 = os.environ['AMS_EMPRESA_1']
EMPRESA_2 = os.environ['AMS_EMPRESA_2']
APP_KEY = os.environ['DROPBOX_APP_KEY']
APP_SECRET = os.environ['DROPBOX_APP_SECRET']
REFRESH_TOKEN = os.environ['DROPBOX_REFRESH_TOKEN']
GH_TOKEN = os.environ['GH_TOKEN_PUSH']
GH_REPO = 'ignaciopacci/tiburon-dashboard'
GOOGLE_CREDENTIALS = os.environ['GOOGLE_CREDENTIALS']
GANANCIAS_FILE_ID = '1b2rxkDVO9ujMurm19gqOd9C1myfDt_b_'
# Excel de costos de pinceles — hoja "PRECIO CERDA GRAMOS REAL" tiene los precios reales actualizados
COSTOS_CERDA_FILE_ID = '1lHIOv2TvbY_6REjYnbzksVXqqucWo_eC'

URL_LOGIN = 'https://apps1.mahonsistemas.com.ar/WebCorporateTiburon/login.aspx'
BASE = 'https://apps1.mahonsistemas.com.ar/WebCorporateTiburon/'

# ───────────────────────────────────────────────────────────
# TABLA DE GRAMOS DE CERDA POR LÍNEA Y MEDIDA
# Extraída de COSTO_DE_PINCELES_NUEVO_-_ANALISIS_2021.xlsx > PRECIO CERDA GRAMOS NUEVO
# ───────────────────────────────────────────────────────────
GRAMOS_CERDA = {
    ('Clasica', '7/1'): 4.0, ('Clasica', '10/1'): 5.0, ('Clasica', '15/1'): 8.0,
    ('Clasica', '20/1'): 10.0, ('Clasica', '25/1'): 15.0, ('Clasica', '30/1'): 18.0,
    ('Clasica', '7/2'): 5.0, ('Clasica', '10/2'): 6.0, ('Clasica', '15/2'): 13.0,
    ('Clasica', '20/2'): 15.0, ('Clasica', '25/2'): 19.0, ('Clasica', '30/2'): 23.0,
    ('Ecology', '7/1'): 4.0, ('Ecology', '10/1'): 5.0, ('Ecology', '15/1'): 8.0,
    ('Ecology', '20/1'): 10.0, ('Ecology', '25/1'): 15.0, ('Ecology', '30/1'): 18.0,
    ('Ecology', '7/2'): 5.0, ('Ecology', '10/2'): 6.0, ('Ecology', '15/2'): 13.0,
    ('Ecology', '20/2'): 15.0, ('Ecology', '25/2'): 19.0, ('Ecology', '30/2'): 23.0,
    ('Profesional', '7/1'): 3.9, ('Profesional', '10/1'): 5.2, ('Profesional', '15/1'): 10.4,
    ('Profesional', '20/1'): 13.0, ('Profesional', '25/1'): 19.5, ('Profesional', '30/1'): 23.4,
    ('Profesional', '7/2'): 6.5, ('Profesional', '10/2'): 7.8, ('Profesional', '15/2'): 14.3,
    ('Profesional', '20/2'): 19.5, ('Profesional', '25/2'): 23.0, ('Profesional', '30/2'): 30.0,
    ('Extra Profesional', '10/2'): 8.0, ('Extra Profesional', '15/2'): 15.0,
    ('Extra Profesional', '20/2'): 20.0, ('Extra Profesional', '25/2'): 26.0, ('Extra Profesional', '30/2'): 32.0,
    ('Extra Profesional', '10/4'): 10.0, ('Extra Profesional', '15/4'): 20.0,
    ('Extra Profesional', '20/4'): 32.0, ('Extra Profesional', '25/4'): 40.0, ('Extra Profesional', '30/4'): 62.0,
    ('Bacota', '10/2'): 8.0, ('Bacota', '15/2'): 13.0, ('Bacota', '20/2'): 22.1,
    ('Bacota', '25/2'): 30.0, ('Bacota', '30/2'): 40.0,
    ('Linea1200', '7/1'): 3.0, ('Linea1200', '10/1'): 4.0, ('Linea1200', '15/1'): 10.0,
    ('Linea1200', '20/1'): 12.0, ('Linea1200', '25/1'): 20.0, ('Linea1200', '30/1'): 23.0,
    ('Linea2400', '7/2'): 5.0, ('Linea2400', '10/2'): 7.0, ('Linea2400', '15/2'): 15.0,
    ('Linea2400', '20/2'): 23.0, ('Linea2400', '25/2'): 26.0, ('Linea2400', '30/2'): 32.0,
    ('Remox', '10/1'): 3.0, ('Remox', '15/1'): 4.0, ('Remox', '20/1'): 6.0,
    ('Remox', '25/1'): 10.0, ('Remox', '30/1'): 11.0,
}

# Precio de cerda CON COLCHÓN (USD/kg) por línea — valores actuales del Excel de costos
PRECIO_CERDA_COLCHON = {
    'Clasica': 6.64, 'Ecology': 6.64, 'Linea1200': 6.64,
    'Profesional': 8.65, 'Extra Profesional': 8.65, 'Bacota': 8.65,
    'Linea2400': 8.65, 'Remox': 8.65,
}

DOLAR_COLCHON_EXCEL = 1500
FACTOR_GASTOS_IMPORT = 1.7  # +70% sobre el precio bruto de cerda

def extraer_codigo_tamano(articulo):
    art = articulo.upper()
    match_codigo = re.search(r'\(([A-Z]+)-', art)
    codigo = match_codigo.group(1) if match_codigo else None
    match_tam = re.search(r'(\d{1,2}/\d)', art)
    tamano = match_tam.group(1) if match_tam else None
    return codigo, tamano

def clasificar_linea_pincel(articulo, codigo):
    art = articulo.upper()
    if 'ECOLOGY' in art:
        return 'Ecology'
    if 'CLASICO' in art or codigo == 'PEC':
        return 'Clasica'
    if 'EXTRA PROF' in art or codigo == 'PLE':
        return 'Extra Profesional'
    if 'PROFESIONAL' in art or codigo == 'PLT':
        return 'Profesional'
    if 'BACOTA' in art or codigo == 'PLB':
        return 'Bacota'
    if 'PINCELETA' in art or codigo == 'PLP':
        return 'Pinceleta'
    if '1200' in art:
        return 'Linea1200'
    if '2400' in art:
        return 'Linea2400'
    if 'REMOX' in art:
        return 'Remox'
    if 'FLORENTINA' in art:
        return 'Florentina'
    return 'Otro'

# ───────────────────────────────────────────────────────────

def get_hora_arg():
    try:
        r = requests.get('https://worldtimeapi.org/api/timezone/America/Argentina/Buenos_Aires', timeout=5)
        dt_str = r.json()['datetime'][:19]
        return datetime.fromisoformat(dt_str)
    except:
        return datetime.now(ARG).replace(tzinfo=None)

def get_rango_mes():
    hoy = get_hora_arg()
    primer_dia = hoy.replace(day=1).strftime('%Y%m%d')
    ultimo = calendar.monthrange(hoy.year, hoy.month)[1]
    ultimo_dia = hoy.replace(day=ultimo).strftime('%Y%m%d')
    return primer_dia, ultimo_dia

def get_url_reporte_costo():
    primer_dia, ultimo_dia = get_rango_mes()
    return f'{BASE}alstinfcompcosto.aspx?{primer_dia},{ultimo_dia},PES,,A,SCR'

def get_url_reporte_ventas():
    primer_dia, ultimo_dia = get_rango_mes()
    # Rubro 001 = Pinceles, orden por Articulo
    return f'{BASE}alstinfestventas.aspx?{primer_dia},{ultimo_dia},PES,0,,0,,001,A,A,SCR'

def login(page, empresa, valor):
    print(f'Entrando como: [{empresa}] (value={valor})')
    page.goto(URL_LOGIN)
    page.wait_for_load_state('networkidle')
    page.wait_for_selector('#vUSUARIOCOD', timeout=15000)
    page.fill('#vUSUARIOCOD', USUARIO)
    page.press('#vUSUARIOCOD', 'Tab')
    page.wait_for_timeout(2000)
    page.evaluate(f'''
        var sel = document.querySelector('#vEMPRESACGO');
        sel.value = '{valor}';
        sel.dispatchEvent(new Event('change', {{bubbles: true}}));
    ''')
    page.wait_for_timeout(1000)
    seleccionado = page.evaluate("document.querySelector('#vEMPRESACGO').value")
    print(f'Empresa seleccionada value: {seleccionado}')
    page.fill('#vUSUARIOPASS', PASSWORD)
    page.wait_for_timeout(500)
    page.evaluate('''
        var btns = document.querySelectorAll("input[type=button], input[type=submit], button");
        for (var i = 0; i < btns.length; i++) {
            if (btns[i].value && btns[i].value.toLowerCase().includes("ingres")) {
                btns[i].click(); break;
            }
        }
    ''')
    page.wait_for_load_state('networkidle')
    page.wait_for_timeout(2000)
    print(f'Login OK: [{empresa}]')

def descargar_reporte(page, context, url_reporte, referer_page='InfCompCosto.aspx'):
    print(f'Descargando: {url_reporte}')
    cookies = context.cookies()
    session = requests.Session()
    for cookie in cookies:
        session.cookies.set(cookie['name'], cookie['value'])
    session.headers.update({
        'User-Agent': 'Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36',
        'Referer': BASE + referer_page,
    })
    response = session.get(url_reporte, allow_redirects=True)
    print(f'Reporte: {len(response.content)} bytes')
    return response.content

def parsear_pdf_costo(pdf_bytes):
    """Parsea el PDF de Comparativo de Costo (formato original, todos los rubros)."""
    datos = []
    rubro_actual = None
    with pdfplumber.open(BytesIO(pdf_bytes)) as pdf:
        for page in pdf.pages:
            texto = page.extract_text()
            if not texto:
                continue
            for linea in texto.split('\n'):
                linea = linea.strip()
                if not linea:
                    continue
                if 'Rubro Pinceles' in linea or 'Pinceles(001)' in linea:
                    rubro_actual = 'Pinceles'
                    continue
                if 'Rubro Accesorios' in linea or 'Accesorios(002)' in linea:
                    rubro_actual = 'Accesorios'
                    continue
                if 'Rubro Rodillos' in linea or 'Rodillos(003)' in linea:
                    rubro_actual = 'Rodillos'
                    continue
                if not rubro_actual:
                    continue
                match = re.search(
                    r'^(.+?\([A-Z0-9\-]+\))\s+([\d\-]+)\s+([\d\.,]+)\s+([\d\.,]+)\s+([\d\.,]+)\s*$',
                    linea
                )
                if match:
                    try:
                        articulo = match.group(1).strip()
                        cantidad = float(match.group(2).replace('.', '').replace(',', '.'))
                        costo_unit = float(match.group(3).replace('.', '').replace(',', '.'))
                        total_costo = float(match.group(4).replace('.', '').replace(',', '.'))
                        total_fac = float(match.group(5).replace('.', '').replace(',', '.'))
                        datos.append({
                            'rubro': rubro_actual,
                            'articulo': articulo,
                            'cantidad': cantidad,
                            'ultimoCosto': costo_unit,
                            'totalCosto': total_costo,
                            'totalFac': total_fac
                        })
                    except:
                        continue
    print(f'Artículos parseados (costo): {len(datos)}')
    return datos

def parsear_pdf_ventas_pinceles(pdf_bytes):
    """
    Parsea el PDF de Estadística de Ventas filtrado a rubro Pinceles.
    Formato esperado por línea: Articulo ... Cantidad ... Precio_unit ... Total
    Devuelve lista de {'articulo': str, 'cantidad': float}
    """
    datos = []
    articulo_actual = None
    cantidad_acumulada = {}
    with pdfplumber.open(BytesIO(pdf_bytes)) as pdf:
        for page in pdf.pages:
            texto = page.extract_text()
            if not texto:
                continue
            for linea in texto.split('\n'):
                linea = linea.strip()
                if not linea:
                    continue
                # Detectar nombre de artículo (contiene paréntesis con código)
                match_articulo = re.search(r'^(Pincel.+?\([A-Z]+-\d+\))', linea)
                if match_articulo:
                    articulo_actual = match_articulo.group(1).strip()
                    continue
                # Detectar línea de venta: fecha, comprobante, cuenta, cantidad, unid, precio, total...
                match_venta = re.search(
                    r'([\d\.,]+)\s+U\s+([\d\.,]+)\s+([\d\.,]+)',
                    linea
                )
                if match_venta and articulo_actual:
                    try:
                        cantidad = float(match_venta.group(1).replace('.', '').replace(',', '.'))
                        cantidad_acumulada[articulo_actual] = cantidad_acumulada.get(articulo_actual, 0) + cantidad
                    except:
                        continue
    for articulo, cantidad in cantidad_acumulada.items():
        datos.append({'articulo': articulo, 'cantidad': cantidad})
    print(f'Artículos de venta parseados (pinceles): {len(datos)}')
    return datos

def calcular_ajuste_cerda(ventas_pinceles, precios_reales_cerda, dolar_real):
    """
    ventas_pinceles: lista de {'articulo':, 'cantidad':}
    precios_reales_cerda: dict {linea: precio_usd_kg_real}
    dolar_real: dólar billete actual a usar para valorizar el ahorro
    Nota: Remox usa cerda nacional (cerda vaca) sin colchón aplicado — su costo
    registrado ya es el real, por lo que no participa de este ajuste.
    """
    kg_por_linea = {}
    for item in ventas_pinceles:
        articulo = item['articulo']
        cantidad = item['cantidad']
        codigo, tamano = extraer_codigo_tamano(articulo)
        linea = clasificar_linea_pincel(articulo, codigo)
        gramos_unit = GRAMOS_CERDA.get((linea, tamano), 0)
        if gramos_unit > 0:
            kg = (cantidad * gramos_unit) / 1000
            kg_por_linea[linea] = kg_por_linea.get(linea, 0) + kg

    ahorro_total = 0
    detalle = {}
    for linea, kg in kg_por_linea.items():
        if linea == 'Remox':
            continue  # sin colchón, no aplica ajuste
        precio_colchon = PRECIO_CERDA_COLCHON.get(linea)
        precio_real = precios_reales_cerda.get(linea) if precios_reales_cerda else None
        if precio_colchon and precio_real and kg > 0:
            diff = precio_colchon - precio_real
            ahorro = kg * diff * FACTOR_GASTOS_IMPORT * DOLAR_COLCHON_EXCEL
            ahorro_total += ahorro
            detalle[linea] = {
                'kg': round(kg, 2),
                'precioColchon': precio_colchon,
                'precioReal': precio_real,
                'ahorro': round(ahorro, 0)
            }

    return round(ahorro_total, 0), detalle

def generar_json(datos, empresa_nombre):
    hoy = get_hora_arg()
    rubros = {}
    for d in datos:
        r = d['rubro']
        if r not in rubros:
            rubros[r] = {'unidades': 0, 'totalFac': 0, 'totalCosto': 0, 'articulos': []}
        rubros[r]['unidades'] += d['cantidad']
        rubros[r]['totalFac'] += d['totalFac']
        rubros[r]['totalCosto'] += d['totalCosto']
        rubros[r]['articulos'].append(d)
    total_fac = sum(r['totalFac'] for r in rubros.values())
    total_costo = sum(r['totalCosto'] for r in rubros.values())
    return {
        'empresa': empresa_nombre,
        'fechaActualizacion': hoy.strftime('%d/%m/%Y %H:%M'),
        'mes': hoy.strftime('%m/%Y'),
        'anio': hoy.year,
        'nroMes': hoy.month,
        'totalFac': total_fac,
        'totalCosto': total_costo,
        'ganancia': total_fac - total_costo,
        'margen': round((total_fac - total_costo) / total_fac * 100, 1) if total_fac else 0,
        'rubros': rubros
    }

def subir_github(contenido_str, path):
    url = f'https://api.github.com/repos/{GH_REPO}/contents/{path}'
    headers = {
        'Authorization': f'token {GH_TOKEN}',
        'Accept': 'application/vnd.github.v3+json'
    }
    r = requests.get(url, headers=headers)
    sha = r.json().get('sha') if r.status_code == 200 else None
    contenido_b64 = base64.b64encode(contenido_str.encode('utf-8')).decode('utf-8')
    data = {
        'message': f'Actualización {datetime.now(ARG).strftime("%d/%m/%Y %H:%M")}',
        'content': contenido_b64
    }
    if sha:
        data['sha'] = sha
    r = requests.put(url, headers=headers, json=data)
    if r.status_code in [200, 201]:
        print(f'✓ GitHub: {path}')
    else:
        raise Exception(f'Error GitHub {r.status_code}: {path}')

def get_dolar_mes(anio, mes):
    tipos = ['bolsa', 'blue'] if anio >= 2019 else ['blue']
    fecha = f'{anio}/{mes:02d}/15'
    for tipo in tipos:
        try:
            url = f'https://api.argentinadatos.com/v1/cotizaciones/dolares/{tipo}/{fecha}'
            r = requests.get(url, timeout=10)
            if r.status_code == 200:
                data = r.json()
                venta = data.get('venta') or data.get('compra')
                if venta:
                    print(f'  Dólar {tipo} {anio}/{mes:02d}: ${venta}')
                    return {'tipo': tipo, 'valor': float(venta)}
        except Exception as e:
            print(f'  ✗ Error dólar {tipo} {anio}/{mes:02d}: {e}')
    print(f'  ✗ Sin dólar para {anio}/{mes:02d}')
    return {'tipo': 'n/d', 'valor': None}

def get_inflacion_mensual():
    try:
        url = 'https://api.argentinadatos.com/v1/finanzas/indices/inflacion'
        r = requests.get(url, timeout=15)
        if r.status_code == 200:
            data = r.json()
            result = {}
            for item in data:
                fecha = item.get('fecha', '')
                valor = item.get('valor')
                if fecha and valor is not None:
                    partes = fecha.split('-')
                    if len(partes) >= 2:
                        anio = int(partes[0])
                        mes = int(partes[1])
                        result[(anio, mes)] = float(valor)
            print(f'  ✓ Inflación: {len(result)} registros mensuales')
            return result
    except Exception as e:
        print(f'  ✗ Error inflación: {e}')
    return {}

def safe_float(val):
    try:
        if val is None or val == '':
            return 0
        return float(str(val).replace('$', '').replace(' ', '').strip())
    except:
        return 0

def safe_int(val):
    try:
        if val is None or val == '':
            return 0
        return int(float(val))
    except:
        return 0

def descargar_excel_drive(file_id, scopes=None):
    creds_dict = json.loads(GOOGLE_CREDENTIALS)
    creds = service_account.Credentials.from_service_account_info(
        creds_dict,
        scopes=scopes or ['https://www.googleapis.com/auth/drive.readonly']
    )
    service = build('drive', 'v3', credentials=creds)
    fh = io.BytesIO()
    request = service.files().get_media(fileId=file_id)
    downloader = MediaIoBaseDownload(fh, request)
    done = False
    while not done:
        _, done = downloader.next_chunk()
    fh.seek(0)
    return openpyxl.load_workbook(fh, data_only=True)

def leer_precios_reales_cerda():
    """
    Lee la hoja 'PRECIO CERDA GRAMOS REAL' del Excel de costos de pinceles en Drive.
    La hoja tiene múltiples secciones, una por línea (LINEA PROFESIONAL, LINEA CLASICO,
    LINEA EXTRA PROFESIONAL, LINEA BACOTA, LINEA 1200, LINEA 2400, LINEA REMOX, etc.)
    Cada sección tiene una tabla: Cerda(medida mm) | Descripcion | Precio/kilo | precio/gramo
    EXCEPTO Remox, cuyo precio puede estar directo en PESOS (no USD) por ser cerda nacional.
    Devuelve: (precios_reales: dict de USD/kg, dolar_billete: float, remox_precio_pesos: float|None)
    """
    # Mapeo de texto de sección -> nombres internos (puede haber varias líneas por sección)
    SECCIONES = {
        'LINEA PROFESIONAL': ['Profesional'],
        'LINEA EXTRA PROFESIONAL': ['Extra Profesional'],
        'LINEA BACOTA': ['Bacota'],
        'LINEA CLASICO': ['Clasica', 'Ecology'],
        'LINEA CLASICA': ['Clasica', 'Ecology'],
        'LINEA ECOLOGY': ['Ecology'],
        'LINEA 1200': ['Linea1200'],
        'LINEA 2400': ['Linea2400'],
        'LINEA REMOX': ['Remox'],
    }

    try:
        wb = descargar_excel_drive(COSTOS_CERDA_FILE_ID)
        if 'PRECIO CERDA GRAMOS REAL' not in wb.sheetnames:
            print('  ⚠ Hoja PRECIO CERDA GRAMOS REAL no encontrada')
            return None, None, None
        ws = wb['PRECIO CERDA GRAMOS REAL']

        precios_reales = {}
        remox_precio_pesos = None
        seccion_actual = None
        primer_precio_de_seccion = None  # reset por cada nueva sección

        for row in ws.iter_rows(values_only=True):
            if not any(v is not None for v in row):
                continue
            primera_raw = row[0]
            primera = str(primera_raw).strip().upper() if primera_raw else ''

            # ¿Es un header de sección?
            seccion_detectada = None
            for clave, _nombres in SECCIONES.items():
                if clave in primera:
                    seccion_detectada = clave
                    break
            if seccion_detectada:
                seccion_actual = seccion_detectada
                primer_precio_de_seccion = None
                continue

            if not seccion_actual:
                continue

            nombres_linea = SECCIONES[seccion_actual]

            # Caso Remox: filas con código numérico (52, 53...) y precio en PESOS (>1000) en columna C
            if seccion_actual == 'LINEA REMOX':
                if isinstance(primera_raw, (int, float)) and remox_precio_pesos is None:
                    precio_col = row[2] if len(row) > 2 else None
                    if isinstance(precio_col, (int, float)) and precio_col > 1000:
                        remox_precio_pesos = round(float(precio_col), 0)
                continue

            # Caso general: fila con medida en mm (ej '51mm') y precio/kilo en USD en columna C
            if primera_raw and 'MM' in primera and primer_precio_de_seccion is None:
                precio_kilo = row[2] if len(row) > 2 else None
                if isinstance(precio_kilo, (int, float)) and 0.5 < precio_kilo < 50:
                    primer_precio_de_seccion = round(float(precio_kilo), 2)
                    for nombre in nombres_linea:
                        if nombre not in precios_reales:
                            precios_reales[nombre] = primer_precio_de_seccion

        if precios_reales:
            print(f'  ✓ Precios reales cerda (USD/kg): {precios_reales}')
        else:
            print('  ⚠ No se pudieron extraer precios reales de la hoja')
        if remox_precio_pesos:
            print(f'  ✓ Precio real Remox (ARS/kg, cerda nacional): ${remox_precio_pesos:,.0f}')

        dolar_info = get_dolar_mes(get_hora_arg().year, get_hora_arg().month)
        dolar_billete = dolar_info['valor']

        return (precios_reales if precios_reales else None), dolar_billete, remox_precio_pesos
    except Exception as e:
        print(f'  ✗ Error leyendo PRECIO CERDA GRAMOS REAL: {e}')
        return None, None, None

def leer_ganancias_drive():
    print('Leyendo Excel de ganancias desde Google Drive...')
    wb = descargar_excel_drive(GANANCIAS_FILE_ID)
    print(f'Hojas: {wb.sheetnames}')

    resultado = {'anios': {}, 'mensual': [], 'meta': {}}

    if 'VENTAS' in wb.sheetnames:
        ws = wb['VENTAS']
        meses_map = {
            'ENERO': 1, 'FEBRERO': 2, 'MARZO': 3, 'ABRIL': 4,
            'MAYO': 5, 'JUNIO': 6, 'JULIO': 7, 'AGOSTO': 8,
            'SEPTIEMBRE': 9, 'OCTUBRE': 10, 'NOVIEMBRE': 11, 'DICIEMBRE': 12
        }
        meses_nombre = {
            1: 'Ene', 2: 'Feb', 3: 'Mar', 4: 'Abr', 5: 'May', 6: 'Jun',
            7: 'Jul', 8: 'Ago', 9: 'Sep', 10: 'Oct', 11: 'Nov', 12: 'Dic'
        }
        anio_actual = None

        for row in ws.iter_rows(values_only=True):
            if not row[0]:
                continue
            celda = str(row[0]).strip().upper()

            if 'AÑO' in celda or 'ANO' in celda:
                partes = celda.split()
                for p in partes:
                    p_clean = p.replace('.0', '')
                    if p_clean.isdigit() and len(p_clean) == 4:
                        anio_actual = int(p_clean)
                        break
                continue

            if celda in ('TOTAL', 'TOAL') and anio_actual:
                meses_anio = [m for m in resultado['mensual'] if m['anio'] == anio_actual]
                if meses_anio:
                    tot_ventas = sum(m['ventas'] for m in meses_anio)
                    tot_costo = sum(m['costo'] for m in meses_anio)
                    tot_gb = sum(m['gananciaBruta'] for m in meses_anio)
                    tot_gastos = sum(m['gastos'] for m in meses_anio)
                    tot_gl = sum(m['gananciaLimpia'] for m in meses_anio)
                    renta_anual = round(tot_gl / tot_ventas * 100, 1) if tot_ventas else None
                    resultado['anios'][anio_actual] = {
                        'anio': anio_actual,
                        'ventas': tot_ventas,
                        'costo': tot_costo,
                        'gananciaBruta': tot_gb,
                        'gastos': tot_gastos,
                        'gananciaLimpia': tot_gl,
                        'rentaAnual': renta_anual,
                    }
                continue

            if anio_actual and celda in meses_map:
                nro_mes = meses_map[celda]
                try:
                    ventas = safe_float(row[1])
                    costo = safe_float(row[2])
                    gb = safe_float(row[3])
                    gastos = safe_float(row[4])
                    gl = safe_float(row[5])
                    if ventas or gb:
                        resultado['mensual'].append({
                            'anio': anio_actual,
                            'mes': nro_mes,
                            'mesNombre': meses_nombre[nro_mes],
                            'ventas': ventas,
                            'costo': costo,
                            'gananciaBruta': gb,
                            'gastos': gastos,
                            'gananciaLimpia': gl,
                        })
                except:
                    continue

    print(f'✓ Excel: {len(resultado["anios"])} años, {len(resultado["mensual"])} registros mensuales')

    print('Obteniendo datos de inflación...')
    inflacion = get_inflacion_mensual()

    hoy = get_hora_arg()
    mes_base = hoy.month
    anio_base = hoy.year
    meses_nb = {1:'Ene',2:'Feb',3:'Mar',4:'Abr',5:'May',6:'Jun',7:'Jul',8:'Ago',9:'Sep',10:'Oct',11:'Nov',12:'Dic'}
    resultado['meta']['mesBase'] = f'{meses_nb.get(mes_base,str(mes_base))} {anio_base}'

    registros_ordenados = sorted(inflacion.keys())
    indice_por_mes = {}
    factor_acum = 1.0
    for (a, m) in registros_ordenados:
        factor_acum *= (1 + inflacion[(a, m)] / 100)
        indice_por_mes[(a, m)] = factor_acum
    factor_hoy = factor_acum

    print('Obteniendo tipo de cambio histórico...')
    for item in resultado['mensual']:
        a = item['anio']
        m = item['mes']
        ventas = item['ventas']
        costo = item['costo']
        gb = item['gananciaBruta']
        gastos = item['gastos']
        gl = item['gananciaLimpia']

        dolar_info = get_dolar_mes(a, m)
        item['dolarTipo'] = dolar_info['tipo']
        item['dolarValor'] = dolar_info['valor']
        if dolar_info['valor'] and dolar_info['valor'] > 0:
            item['ventasUsd'] = round(ventas / dolar_info['valor'], 0)
            item['costoUsd'] = round(costo / dolar_info['valor'], 0)
            item['gananciaBrutaUsd'] = round(gb / dolar_info['valor'], 0)
            item['gastosUsd'] = round(gastos / dolar_info['valor'], 0)
            item['gananciaLimpiaUsd'] = round(gl / dolar_info['valor'], 0)
        else:
            item['ventasUsd'] = None
            item['costoUsd'] = None
            item['gananciaBrutaUsd'] = None
            item['gastosUsd'] = None
            item['gananciaLimpiaUsd'] = None

        factor_origen = indice_por_mes.get((a, m))
        if factor_origen and factor_hoy:
            fa = factor_hoy / factor_origen
            item['factorInflacion'] = round(fa, 4)
            item['ventasConstante'] = round(ventas * fa, 0)
            item['costoConstante'] = round(costo * fa, 0)
            item['gananciaBrutaConstante'] = round(gb * fa, 0)
            item['gastosConstante'] = round(gastos * fa, 0)
            item['gananciaLimpiaConstante'] = round(gl * fa, 0)
        else:
            item['factorInflacion'] = None
            item['ventasConstante'] = None
            item['costoConstante'] = None
            item['gananciaBrutaConstante'] = None
            item['gastosConstante'] = None
            item['gananciaLimpiaConstante'] = None

        time.sleep(0.3)

    for anio_key, item in resultado['anios'].items():
        a = int(anio_key) if not isinstance(anio_key, int) else anio_key
        ventas = item['ventas']
        costo = item['costo']
        gb = item['gananciaBruta']
        gastos = item['gastos']
        gl = item['gananciaLimpia']

        dolar_info = get_dolar_mes(a, 12)
        item['dolarTipo'] = dolar_info['tipo']
        item['dolarValor'] = dolar_info['valor']
        if dolar_info['valor'] and dolar_info['valor'] > 0 and gb:
            item['ventasUsd'] = round(ventas / dolar_info['valor'], 0)
            item['costoUsd'] = round(costo / dolar_info['valor'], 0)
            item['gananciaBrutaUsd'] = round(gb / dolar_info['valor'], 0)
            item['gastosUsd'] = round(gastos / dolar_info['valor'], 0)
            item['gananciaLimpiaUsd'] = round(gl / dolar_info['valor'], 0)
        else:
            item['ventasUsd'] = None
            item['costoUsd'] = None
            item['gananciaBrutaUsd'] = None
            item['gastosUsd'] = None
            item['gananciaLimpiaUsd'] = None

        factor_origen = indice_por_mes.get((a, 12))
        if factor_origen and factor_hoy and gb:
            fa = factor_hoy / factor_origen
            item['factorInflacion'] = round(fa, 4)
            item['ventasConstante'] = round(ventas * fa, 0)
            item['costoConstante'] = round(costo * fa, 0)
            item['gananciaBrutaConstante'] = round(gb * fa, 0)
            item['gastosConstante'] = round(gastos * fa, 0)
            item['gananciaLimpiaConstante'] = round(gl * fa, 0)
        else:
            item['factorInflacion'] = None
            item['ventasConstante'] = None
            item['costoConstante'] = None
            item['gananciaBrutaConstante'] = None
            item['gastosConstante'] = None
            item['gananciaLimpiaConstante'] = None

        time.sleep(0.3)

    print(f'✓ Ganancias completas: {len(resultado["anios"])} años, {len(resultado["mensual"])} meses')
    return resultado

def main():
    hoy = get_hora_arg()
    mes_key = hoy.strftime('%Y-%m')
    print(f'Hora Argentina: {hoy.strftime("%d/%m/%Y %H:%M")}')

    empresas = [
        (EMPRESA_1, '1'),
        (EMPRESA_2, '2'),
    ]

    resultados = {}
    ventas_pinceles_todas = []  # acumula ventas de pinceles de ambas empresas

    for empresa, valor in empresas:
        with sync_playwright() as p:
            browser = p.chromium.launch(headless=True)
            context = browser.new_context(accept_downloads=True)
            page = context.new_page()
            try:
                login(page, empresa, valor)

                # Reporte de costo (como antes)
                pdf_costo = descargar_reporte(page, context, get_url_reporte_costo(), 'InfCompCosto.aspx')
                datos = parsear_pdf_costo(pdf_costo)
                resultado = generar_json(datos, empresa)
                nombre = empresa.replace(' ', '_').replace('.', '')
                resultados[nombre] = resultado
                print(f'✓ {empresa} — {len(datos)} artículos (costo)')

                # Reporte de ventas de pinceles (para ajuste de cerda)
                try:
                    pdf_ventas = descargar_reporte(page, context, get_url_reporte_ventas(), 'InfEstVentas.aspx')
                    ventas_pinc = parsear_pdf_ventas_pinceles(pdf_ventas)
                    ventas_pinceles_todas.extend(ventas_pinc)
                    print(f'✓ {empresa} — {len(ventas_pinc)} artículos (ventas pinceles)')
                except Exception as e:
                    print(f'✗ Error reporte ventas pinceles {empresa}: {e}')

            except Exception as e:
                print(f'✗ Error: {e}')
                raise
            finally:
                context.close()
                browser.close()
        time.sleep(3)

    for nombre, resultado in resultados.items():
        contenido = json.dumps(resultado, ensure_ascii=False, indent=2)
        subir_github(contenido, f'data/{nombre}.json')
        subir_github(contenido, f'data/historico/{mes_key}_{nombre}.json')

    # Ajuste real de cerda
    try:
        print('Calculando ajuste real de cerda...')
        precios_reales, dolar_billete, remox_precio_real_pesos = leer_precios_reales_cerda()
        if precios_reales and ventas_pinceles_todas:
            ahorro_total, detalle = calcular_ajuste_cerda(ventas_pinceles_todas, precios_reales, dolar_billete)
            ajuste_cerda = {
                'mes': hoy.strftime('%m/%Y'),
                'fechaCalculo': hoy.strftime('%d/%m/%Y %H:%M'),
                'ahorroTotal': ahorro_total,
                'detallePorLinea': detalle,
                'dolarBilleteUsado': dolar_billete,
            }
            subir_github(
                json.dumps(ajuste_cerda, ensure_ascii=False, indent=2),
                'data/ajuste_cerda_mes_actual.json'
            )
            print(f'✓ Ajuste cerda calculado: ${ahorro_total:,.0f}')
        else:
            print('  ⚠ No se pudo calcular ajuste de cerda (faltan precios reales o ventas)')
    except Exception as e:
        print(f'✗ Error calculando ajuste de cerda: {e}')

    # Ganancias históricas
    try:
        ganancias = leer_ganancias_drive()
        subir_github(
            json.dumps(ganancias, ensure_ascii=False, indent=2),
            'data/ganancias_historicas.json'
        )
    except Exception as e:
        print(f'✗ Error leyendo ganancias Drive: {e}')

    print('Completo:', hoy.strftime('%d/%m/%Y %H:%M'))

if __name__ == '__main__':
    main()
