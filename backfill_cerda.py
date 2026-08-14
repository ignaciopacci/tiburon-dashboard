"""
Script de backfill — calcula el ajuste real de cerda para meses históricos.
Descarga el reporte de Estadística de Ventas de cada mes especificado,
calcula el ajuste de cerda y lo guarda en data/ajuste_cerda_historico.json.

Uso: python backfill_cerda.py
Configurar MESES_BACKFILL con los meses a procesar.
"""
import os
import json
import re
import requests
import base64
import io
import time
from datetime import datetime, timedelta, timezone
import calendar
import pdfplumber
from io import BytesIO
from playwright.sync_api import sync_playwright
from google.oauth2 import service_account
from googleapiclient.discovery import build
from googleapiclient.http import MediaIoBaseDownload
import openpyxl

ARG = timezone(timedelta(hours=-3))

USUARIO = os.environ['AMS_USUARIO']
PASSWORD = os.environ['AMS_PASSWORD']
EMPRESA_1 = os.environ['AMS_EMPRESA_1']
EMPRESA_2 = os.environ['AMS_EMPRESA_2']
GH_TOKEN = os.environ['GH_TOKEN_PUSH']
GH_REPO = 'ignaciopacci/tiburon-dashboard'
GOOGLE_CREDENTIALS = os.environ['GOOGLE_CREDENTIALS']
COSTOS_CERDA_FILE_ID = '1lHIOv2TvbY_6REjYnbzksVXqqucWo_eC'

URL_LOGIN = 'https://apps1.mahonsistemas.com.ar/WebCorporateTiburon/login.aspx'
BASE = 'https://apps1.mahonsistemas.com.ar/WebCorporateTiburon/'

# ── Meses a procesar ──────────────────────────────────────
# Formato: (año, mes) — agregar todos los meses que faltan
MESES_BACKFILL = [
    (2026, 1),  # Enero
    (2026, 2),  # Febrero
    (2026, 3),  # Marzo
    (2026, 4),  # Abril
    (2026, 5),  # Mayo
    (2026, 6),  # Junio
]

# ── Tabla de gramos de cerda (igual que en scraper.py) ────
GRAMOS_CERDA = {
    ('Clasica', '7/1'): 4.0, ('Clasica', '10/1'): 5.0, ('Clasica', '15/1'): 8.0,
    ('Clasica', '20/1'): 10.0, ('Clasica', '25/1'): 15.0, ('Clasica', '30/1'): 18.0,
    ('Clasica', '7/2'): 5.0, ('Clasica', '10/2'): 6.0, ('Clasica', '15/2'): 13.0,
    ('Clasica', '20/2'): 15.0, ('Clasica', '25/2'): 19.0, ('Clasica', '30/2'): 23.0,
    ('Ecology', '7/1'): 4.0, ('Ecology', '10/1'): 5.0, ('Ecology', '15/1'): 8.0,
    ('Ecology', '20/1'): 10.0, ('Ecology', '25/1'): 15.0, ('Ecology', '30/1'): 18.0,
    ('Ecology', '7/2'): 5.0, ('Ecology', '10/2'): 6.0, ('Ecology', '15/2'): 13.0,
    ('Ecology', '20/2'): 15.0, ('Ecology', '25/2'): 19.0, ('Ecology', '30/2'): 23.0,
    ('Profesional', '7/1'): 3.9, ('Profesional', '10/1'): 5.2, ('Profesional', '15/1'): 10.4,
    ('Profesional', '20/1'): 13.0, ('Profesional', '25/1'): 19.5, ('Profesional', '30/1'): 23.4,
    ('Profesional', '7/2'): 6.5, ('Profesional', '10/2'): 7.8, ('Profesional', '15/2'): 14.3,
    ('Profesional', '20/2'): 19.5, ('Profesional', '25/2'): 23.0, ('Profesional', '30/2'): 30.0,
    ('Extra Profesional', '10/2'): 8.0, ('Extra Profesional', '15/2'): 15.0,
    ('Extra Profesional', '20/2'): 20.0, ('Extra Profesional', '25/2'): 26.0, ('Extra Profesional', '30/2'): 32.0,
    ('Extra Profesional', '10/4'): 10.0, ('Extra Profesional', '15/4'): 20.0,
    ('Extra Profesional', '20/4'): 32.0, ('Extra Profesional', '25/4'): 40.0, ('Extra Profesional', '30/4'): 62.0,
    ('Bacota', '10/2'): 8.0, ('Bacota', '15/2'): 13.0, ('Bacota', '20/2'): 22.1,
    ('Bacota', '25/2'): 30.0, ('Bacota', '30/2'): 40.0,
    ('Linea1200', '7/1'): 3.0, ('Linea1200', '10/1'): 4.0, ('Linea1200', '15/1'): 10.0,
    ('Linea1200', '20/1'): 12.0, ('Linea1200', '25/1'): 20.0, ('Linea1200', '30/1'): 23.0,
    ('Linea2400', '7/2'): 5.0, ('Linea2400', '10/2'): 7.0, ('Linea2400', '15/2'): 15.0,
    ('Linea2400', '20/2'): 23.0, ('Linea2400', '25/2'): 26.0, ('Linea2400', '30/2'): 32.0,
    ('Remox', '10/1'): 3.0, ('Remox', '15/1'): 4.0, ('Remox', '20/1'): 6.0,
    ('Remox', '25/1'): 10.0, ('Remox', '30/1'): 11.0,
}

PRECIO_CERDA_COLCHON = {
    'Clasica': 6.64, 'Ecology': 6.64, 'Linea1200': 6.64,
    'Profesional': 8.65, 'Extra Profesional': 8.65, 'Bacota': 8.65,
    'Linea2400': 8.65, 'Remox': 8.65,
}

FACTOR_GASTOS_IMPORT = 1.7

def extraer_codigo_tamano(articulo):
    art = articulo.upper()
    match_codigo = re.search(r'\(([A-Z]+)-', art)
    codigo = match_codigo.group(1) if match_codigo else None
    match_tam = re.search(r'(\d{1,2}/\d)', art)
    tamano = match_tam.group(1) if match_tam else None
    return codigo, tamano

def clasificar_linea_pincel(articulo, codigo):
    art = articulo.upper()
    if 'ECOLOGY' in art: return 'Ecology'
    if 'CLASICO' in art or codigo == 'PEC': return 'Clasica'
    if 'EXTRA PROF' in art or codigo == 'PLE': return 'Extra Profesional'
    if 'PROFESIONAL' in art or codigo == 'PLT': return 'Profesional'
    if 'BACOTA' in art or codigo == 'PLB': return 'Bacota'
    if 'PINCELETA' in art or codigo == 'PLP': return 'Pinceleta'
    if '1200' in art: return 'Linea1200'
    if '2400' in art: return 'Linea2400'
    if 'REMOX' in art: return 'Remox'
    return 'Otro'

def login(page, valor):
    page.goto(URL_LOGIN)
    page.wait_for_load_state('networkidle')
    page.wait_for_selector('#vUSUARIOCOD', timeout=15000)
    page.fill('#vUSUARIOCOD', USUARIO)
    page.press('#vUSUARIOCOD', 'Tab')
    page.wait_for_timeout(2000)
    page.evaluate(f'''
        var sel = document.querySelector('#vEMPRESACGO');
        sel.value = '{valor}';
        sel.dispatchEvent(new Event('change', {{bubbles: true}}));
    ''')
    page.wait_for_timeout(1000)
    page.fill('#vUSUARIOPASS', PASSWORD)
    page.wait_for_timeout(500)
    page.evaluate('''
        var btns = document.querySelectorAll("input[type=button], input[type=submit], button");
        for (var i = 0; i < btns.length; i++) {
            if (btns[i].value && btns[i].value.toLowerCase().includes("ingres")) {
                btns[i].click(); break;
            }
        }
    ''')
    page.wait_for_load_state('networkidle')
    page.wait_for_timeout(2000)

def descargar_pdf_ventas(page, context, anio, mes):
    primer_dia = f'{anio}{mes:02d}01'
    ultimo = calendar.monthrange(anio, mes)[1]
    ultimo_dia = f'{anio}{mes:02d}{ultimo:02d}'
    url = f'{BASE}alstinfestventas.aspx?{primer_dia},{ultimo_dia},PES,0,,0,,001,A,A,SCR'
    cookies = context.cookies()
    session = requests.Session()
    for cookie in cookies:
        session.cookies.set(cookie['name'], cookie['value'])
    session.headers.update({'User-Agent': 'Mozilla/5.0', 'Referer': BASE + 'InfEstVentas.aspx'})
    response = session.get(url, allow_redirects=True)
    return response.content

def parsear_pdf_ventas(pdf_bytes):
    cantidad_acumulada = {}
    articulo_actual = None
    with pdfplumber.open(BytesIO(pdf_bytes)) as pdf:
        for page in pdf.pages:
            texto = page.extract_text()
            if not texto:
                continue
            for linea in texto.split('\n'):
                linea = linea.strip()
                if not linea:
                    continue
                match_articulo = re.search(r'^(Pincel.+?\([A-Z]+-\d+\))', linea)
                if match_articulo:
                    articulo_actual = match_articulo.group(1).strip()
                    continue
                match_venta = re.search(r'\)\s*([\d]{1,4}(?:[.,]\d{1,2})?)\s+U\s+', linea)
                if match_venta and articulo_actual:
                    try:
                        cantidad = float(match_venta.group(1).replace('.', '').replace(',', '.'))
                        if cantidad > 0:
                            cantidad_acumulada[articulo_actual] = cantidad_acumulada.get(articulo_actual, 0) + cantidad
                    except:
                        continue
    return [{'articulo': a, 'cantidad': c} for a, c in cantidad_acumulada.items()]

def leer_precios_reales_cerda():
    creds_dict = json.loads(GOOGLE_CREDENTIALS)
    creds = service_account.Credentials.from_service_account_info(
        creds_dict, scopes=['https://www.googleapis.com/auth/drive.readonly']
    )
    service = build('drive', 'v3', credentials=creds)
    fh = io.BytesIO()
    request = service.files().get_media(fileId=COSTOS_CERDA_FILE_ID)
    from googleapiclient.http import MediaIoBaseDownload
    downloader = MediaIoBaseDownload(fh, request)
    done = False
    while not done:
        _, done = downloader.next_chunk()
    fh.seek(0)
    wb = openpyxl.load_workbook(fh, data_only=True)

    # Leer dólar colchón de I1
    dolar_colchon = 1530
    if 'PRECIO CERDA GRAMOS NUEVO' in wb.sheetnames:
        ws_nuevo = wb['PRECIO CERDA GRAMOS NUEVO']
        val = ws_nuevo['I1'].value
        if isinstance(val, (int, float)) and val > 100:
            dolar_colchon = float(val)

    SECCIONES = {
        'LINEA PROFESIONAL': ['Profesional'],
        'EXTRA PROFESIONAL': ['Extra Profesional'],
        'BACOTA': ['Bacota'],
        'LINEA CLASICO': ['Clasica', 'Ecology'],
        'LINEA CLASICA': ['Clasica', 'Ecology'],
        'LINEA 1200 Y 2400': ['Linea1200', 'Linea2400'],
        'LINEA 1200': ['Linea1200'],
        'LINEA 2400': ['Linea2400'],
        'LINEA REMOX': ['Remox'],
    }
    precios_reales = {}
    seccion_actual = None
    primer_precio = None
    ws = wb['PRECIO CERDA GRAMOS REAL']
    for row in ws.iter_rows(values_only=True):
        if not any(v is not None for v in row):
            continue
        primera_raw = row[0]
        primera = str(primera_raw).strip().upper() if primera_raw else ''
        seccion_detectada = None
        if primera and len(primera) < 40:
            for clave, _ in sorted(SECCIONES.items(), key=lambda x: -len(x[0])):
                if clave in primera:
                    seccion_detectada = clave
                    break
        if seccion_detectada:
            seccion_actual = seccion_detectada
            primer_precio = None
            continue
        if not seccion_actual:
            continue
        if seccion_actual == 'LINEA REMOX':
            continue
        if primera_raw and 'MM' in primera and primer_precio is None:
            precio_kilo = row[2] if len(row) > 2 else None
            if isinstance(precio_kilo, (int, float)) and 0.5 < precio_kilo < 50:
                primer_precio = round(float(precio_kilo), 2)
                for nombre in SECCIONES[seccion_actual]:
                    if nombre not in precios_reales:
                        precios_reales[nombre] = primer_precio

    print(f'  Precios reales: {precios_reales}')
    print(f'  Dólar colchón: ${dolar_colchon}')
    return precios_reales, dolar_colchon

def get_dolar_mes(anio, mes):
    tipos = ['bolsa', 'blue'] if anio >= 2019 else ['blue']
    for tipo in tipos:
        for dia in range(28, 0, -1):
            try:
                url = f'https://api.argentinadatos.com/v1/cotizaciones/dolares/{tipo}/{anio}/{mes:02d}/{dia:02d}'
                r = requests.get(url, timeout=10)
                if r.status_code == 200:
                    data = r.json()
                    venta = data.get('venta') or data.get('compra')
                    if venta:
                        return float(venta)
            except:
                continue
    return None

def calcular_ajuste(ventas_pinceles, precios_reales, dolar_colchon):
    kg_por_linea = {}
    for item in ventas_pinceles:
        codigo, tamano = extraer_codigo_tamano(item['articulo'])
        linea = clasificar_linea_pincel(item['articulo'], codigo)
        gramos = GRAMOS_CERDA.get((linea, tamano), 0)
        if gramos > 0:
            kg = (item['cantidad'] * gramos) / 1000
            kg_por_linea[linea] = kg_por_linea.get(linea, 0) + kg

    ahorro_total = 0
    detalle = {}
    for linea, kg in kg_por_linea.items():
        if linea == 'Remox':
            continue
        precio_colchon = PRECIO_CERDA_COLCHON.get(linea)
        precio_real = precios_reales.get(linea)
        if precio_colchon and precio_real and kg > 0:
            diff = precio_colchon - precio_real
            ahorro = kg * diff * FACTOR_GASTOS_IMPORT * dolar_colchon
            ahorro_total += ahorro
            detalle[linea] = {'kg': round(kg, 2), 'precioColchon': precio_colchon, 'precioReal': precio_real, 'ahorro': round(ahorro, 0)}

    return round(ahorro_total, 0), detalle

def subir_github(contenido_str, path):
    url = f'https://api.github.com/repos/{GH_REPO}/contents/{path}'
    headers = {'Authorization': f'token {GH_TOKEN}', 'Accept': 'application/vnd.github.v3+json'}
    r = requests.get(url, headers=headers)
    sha = r.json().get('sha') if r.status_code == 200 else None
    contenido_b64 = base64.b64encode(contenido_str.encode('utf-8')).decode('utf-8')
    data = {'message': f'Backfill ajuste cerda {datetime.now().strftime("%d/%m/%Y %H:%M")}', 'content': contenido_b64}
    if sha:
        data['sha'] = sha
    r = requests.put(url, headers=headers, json=data)
    if r.status_code in [200, 201]:
        print(f'  ✓ GitHub: {path}')
    else:
        raise Exception(f'Error GitHub {r.status_code}: {path}')

def main():
    print('=== BACKFILL AJUSTE CERDA ===')

    # Leer precios reales de cerda
    precios_reales, dolar_colchon = leer_precios_reales_cerda()

    # Cargar histórico existente
    url_hist = f'https://api.github.com/repos/{GH_REPO}/contents/data/ajuste_cerda_historico.json'
    headers_gh = {'Authorization': f'token {GH_TOKEN}', 'Accept': 'application/vnd.github.v3+json'}
    r_hist = requests.get(url_hist, headers=headers_gh)
    historico = json.loads(base64.b64decode(r_hist.json()['content']).decode('utf-8')) if r_hist.status_code == 200 else {}
    print(f'Histórico actual: {len(historico)} meses')

    empresas = [(EMPRESA_1, '1'), (EMPRESA_2, '2')]

    for anio, mes in MESES_BACKFILL:
        mes_key = f'{anio}-{mes:02d}'
        print(f'\n--- Procesando {mes_key} ---')

        # Si ya existe y está cerrado, saltar
        if mes_key in historico and historico[mes_key].get('cerrado'):
            print(f'  Ya procesado, saltando.')
            continue

        ventas_todas = []
        for empresa, valor in empresas:
            with sync_playwright() as p:
                browser = p.chromium.launch(headless=True)
                context = browser.new_context()
                page = context.new_page()
                try:
                    login(page, valor)
                    pdf = descargar_pdf_ventas(page, context, anio, mes)
                    ventas = parsear_pdf_ventas(pdf)
                    ventas_todas.extend(ventas)
                    print(f'  {empresa}: {len(ventas)} artículos')
                except Exception as e:
                    print(f'  Error {empresa}: {e}')
                finally:
                    context.close()
                    browser.close()
            time.sleep(2)

        if not ventas_todas:
            print(f'  Sin ventas, saltando.')
            continue

        dolar = get_dolar_mes(anio, mes)
        ahorro, detalle = calcular_ajuste(ventas_todas, precios_reales, dolar_colchon)

        historico[mes_key] = {
            'mes': f'{mes:02d}/{anio}',
            'ahorroTotal': ahorro,
            'detallePorLinea': detalle,
            'dolarColchon': dolar_colchon,
            'fechaCalculo': datetime.now().strftime('%d/%m/%Y %H:%M'),
            'cerrado': True,
        }
        print(f'  Ajuste calculado: ${ahorro:,.0f}')

    subir_github(json.dumps(historico, ensure_ascii=False, indent=2), 'data/ajuste_cerda_historico.json')
    print(f'\n✓ Histórico actualizado: {len(historico)} meses')

if __name__ == '__main__':
    main()
